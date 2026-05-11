import openpyxl
from openpyxl.chart import PieChart, Reference
from db import fetch_all
from datetime import datetime
import os

def generate_report_for_supplier(supplier_id):
    # Получаем данные о поставщике
    supplier = fetch_all("SELECT name FROM suppliers WHERE supplier_id = %s", (supplier_id,))
    if not supplier:
        return
    supplier_name = supplier[0]['name']

    # Заказы поставщика (не отменённые)
    orders = fetch_all("""
        SELECT order_id, order_date, status, total_amount
        FROM orders
        WHERE supplier_id = %s AND status != 'Отменен'
    """, (supplier_id,))

    if not orders:
        return

    total_amount = sum(o['total_amount'] for o in orders)
    paid_amount = 0
    all_items = []

    for order in orders:
        # Платежи по заказу
        payments = fetch_all("SELECT COALESCE(SUM(amount),0) as paid FROM payments WHERE order_id = %s", (order['order_id'],))
        paid = payments[0]['paid']
        paid_amount += paid

        # Товары в заказе
        items = fetch_all("""
            SELECT p.name as product, oi.quantity, oi.unit_price, oi.amount
            FROM order_items oi JOIN products p ON oi.product_id = p.product_id
            WHERE oi.order_id = %s
        """, (order['order_id'],))
        all_items.extend(items)

    unpaid = total_amount - paid_amount

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    ws['A1'] = f"Материальный отчет по поставщику: {supplier_name} (ID: {supplier_id})"
    ws['A3'] = f"Общая сумма заказов: {total_amount:.2f} руб."
    ws['A4'] = f"Оплачено: {paid_amount:.2f} руб. ({paid_amount/total_amount*100:.1f}%)"
    ws['A5'] = f"Остаток: {unpaid:.2f} руб. ({unpaid/total_amount*100:.1f}%)"

    ws['A7'] = "Тип"
    ws['B7'] = "Сумма"
    ws['A8'] = "Оплачено"
    ws['B8'] = paid_amount
    ws['A9'] = "Остаток"
    ws['B9'] = unpaid

    # Диаграмма (круговая)
    chart = PieChart()
    chart.title = "Оплата заказов"
    data = Reference(ws, min_col=2, min_row=7, max_row=9)
    labels = Reference(ws, min_col=1, min_row=8, max_row=9)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    ws.add_chart(chart, "D7")

    ws['A11'] = "Товар"
    ws['B11'] = "Количество"
    ws['C11'] = "Цена за ед."
    ws['D11'] = "Сумма"
    for i, item in enumerate(all_items, start=12):
        ws[f'A{i}'] = item['product']
        ws[f'B{i}'] = item['quantity']
        ws[f'C{i}'] = item['unit_price']
        ws[f'D{i}'] = item['amount']

    # Автоширина
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len+2

    # Сохраняем на рабочий стол + открытие
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    filename = f"Отчет_{supplier_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(desktop, filename)
    wb.save(filepath)
    os.startfile(filepath)