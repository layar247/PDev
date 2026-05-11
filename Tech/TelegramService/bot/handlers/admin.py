from aiogram import Router, F
from aiogram.types import Message, CallbackQuery
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from database.models import User, Topic
from database.utils import async_session
from keyboards.admin import admin_menu_kb, cancel_kb
from keyboards.user import main_menu_kb
from sqlalchemy.dialects.mysql import DECIMAL
from texts import ADMIN_START_TEXT, ALLOWED_DIRECTIONS, DIRECTION_WEIGHTS
from sqlalchemy import select
from aiogram import Bot
from sqlalchemy.exc import IntegrityError
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import Filter
from sqlalchemy import delete
from decimal import Decimal
import os
from openpyxl import Workbook,load_workbook

ADMIN_IDS = list(map(int, os.getenv("ADMINS", "").split(",")))

router = Router()

ALLOWED_DIFFICULTIES = ["easy", "medium", "hard"]
ALLOWED_DIRECTIONS = ["ai", "data", "web"]
REQUIRED_WEIGHTS_COUNT = 8


class AddTopic(StatesGroup):
    waiting_for_title = State()
    waiting_for_description = State()
    waiting_for_difficulty = State()
    waiting_for_direction = State()
    waiting_for_weights = State()


class AdminAdd(StatesGroup):
    waiting_for_telegram_id = State()


class IsAdminFilter(Filter):
    async def __call__(self, message: Message) -> bool:
        async with async_session() as session:
            result = await session.execute(
                select(User).where(
                    User.telegram_id == message.from_user.id,
                    User.is_admin == True
                )
            )
            user = result.scalar_one_or_none()
            return user is not None


class AdminRemove(StatesGroup):
    waiting_for_telegram_id = State()


@router.message(Command("admin"))
async def admin_start(message: Message):
    user_id = message.from_user.id

    async with async_session() as session:
        result = await session.execute(
            select(User).where(User.telegram_id == user_id)
        )
        user = result.scalar_one_or_none()

        if user is None and user_id in ADMIN_IDS:
            try:
                chat = await message.bot.get_chat(user_id)
                user = User(
                    telegram_id=user_id,
                    full_name=chat.full_name,
                    username=chat.username,
                    is_admin=True
                )
            except Exception as e:
                print(f"Ошибка получения данных пользователя: {e}")
                user = User(
                    telegram_id=user_id,
                    full_name=f"Admin {user_id}",
                    is_admin=True
                )

            session.add(user)
            await session.commit()

        if user and user.is_admin:
            await message.answer(ADMIN_START_TEXT, reply_markup=admin_menu_kb())
        else:
            await message.answer("🔒 У вас нет прав администратора.")


@router.message(F.text.lower() == "отмена", IsAdminFilter())
async def cancel_handler(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Действие отменено.", reply_markup=admin_menu_kb())


@router.message(F.text == "Удалить администратора")
async def start_remove_admin(message: Message, state: FSMContext):
    # Проверяем, что это владелец
    if message.from_user.id != 717974670:
        await message.answer("❌ У вас нет прав на удаление администраторов.")
        return

    await message.answer(
        "Введите Telegram ID пользователя, которого нужно лишить прав администратора:",
        reply_markup=cancel_kb()
    )
    await state.set_state(AdminRemove.waiting_for_telegram_id)

from aiogram.types import BufferedInputFile
import io
from openpyxl import load_workbook

@router.message(F.text == "Импорт тем из Excel", IsAdminFilter())
async def start_import_topics(message: Message, state: FSMContext):
    await message.answer("Пришлите Excel-файл (.xlsx) с колонками:\n"
                         "Название | Описание | Сложность | Направление | Веса (через пробел)")
    await state.set_state("waiting_for_excel")

@router.message(F.document, IsAdminFilter())
async def handle_excel_upload(message: Message, state: FSMContext, bot: Bot):
    current_state = await state.get_state()
    if current_state != "waiting_for_excel":
        return

    if not message.document.file_name.endswith('.xlsx'):
        await message.answer("❌ Пришлите файл с расширением .xlsx")
        return

    file_data = await bot.download(message.document)
    wb = load_workbook(io.BytesIO(file_data.read()))
    ws = wb.active

    added = 0
    skipped = 0
    async with async_session() as session:
        for row in ws.iter_rows(min_row=2, values_only=True):  # первая строка - заголовки
            title, description, difficulty, direction, weights_str = row[:5]
            if not title:
                continue
            # Парсим веса
            try:
                cleaned = weights_str.replace(',', ' ')
                weights = [float(x) for x in cleaned.split()]
            except (ValueError, AttributeError):
                skipped += 1
                continue

            if len(weights) != REQUIRED_WEIGHTS_COUNT:
                skipped += 1
                continue

            topic = Topic(
                title=title.strip(),
                description=description.strip() if description else "",
                difficulty=difficulty.strip().lower() if difficulty else "easy",
                direction=direction.strip().lower() if direction else "web",
                weights=weights
            )
            session.add(topic)
            added += 1
        await session.commit()

    await message.answer(f"✅ Импорт завершён. Добавлено: {added}, пропущено: {skipped}.")
    await state.clear()
@router.message(AdminRemove.waiting_for_telegram_id)
async def process_remove_admin(message: Message, state: FSMContext):
    if message.from_user.id != 717974670:
        await message.answer("❌ Нет прав.")
        await state.clear()
        return

    try:
        telegram_id = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Отменено. Введите числовой Telegram ID или нажмите 'Отмена'.")
        await state.clear()
        return

    if telegram_id == 717974670:
        await message.answer("❌ Нельзя удалить владельца бота.")
        await state.clear()
        return

    async with async_session() as session:
        result = await session.execute(select(User).where(User.telegram_id == telegram_id))
        user = result.scalar_one_or_none()
        if not user:
            await message.answer(f"❌ Пользователь с ID {telegram_id} не найден.")
            await state.clear()
            return
        if not user.is_admin:
            await message.answer(f"ℹ️ Пользователь {telegram_id} уже не администратор.")
            await state.clear()
            return

        admins_result = await session.execute(select(User).where(User.is_admin == True))
        admins = admins_result.scalars().all()
        if len(admins) <= 1:
            await message.answer("❌ Нельзя удалить последнего администратора.")
            await state.clear()
            return

        user.is_admin = False
        await session.commit()
        await message.answer(
            f"✅ Пользователь <b>{user.full_name or user.telegram_id}</b> (ID: {telegram_id}) больше не администратор.",
            parse_mode="HTML",
            reply_markup=admin_menu_kb()
        )
    await state.clear()


@router.message(F.text == "Список администраторов", IsAdminFilter())
async def show_admins(message: Message, state: FSMContext):
    await state.clear()
    async with async_session() as session:
        result = await session.execute(select(User).where(User.is_admin == True))
        admins = result.scalars().all()

        if not admins:
            await message.answer("Список администраторов пуст.")
            return

        lines = []
        for admin in admins:
            lines.append(f"🆔 <code>{admin.telegram_id}</code> — {admin.full_name or 'Неизвестно'}")

        await message.answer("👑 Список администраторов:\n" + "\n".join(lines), parse_mode="HTML")


@router.message(F.text == "Добавить администратора", IsAdminFilter())
async def start_adding_admin(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Введите Telegram ID пользователя...", reply_markup=cancel_kb())
    await state.set_state(AdminAdd.waiting_for_telegram_id)

from typing import Optional

async def get_user_id_by_username(username: str, bot: Bot) -> Optional[int]:
    if username.startswith('@'):
        username = username[1:]
    try:
        chat = await bot.get_chat(username)
        return chat.id
    except Exception as e:
        print(f"Ошибка получения ID для {username}: {e}")
        return None


@router.message(AdminAdd.waiting_for_telegram_id)
async def process_admin_id(message: Message, state: FSMContext, bot: Bot):
    input_text = message.text.strip()
    telegram_id = None

    if input_text.startswith('@'):
        # Пользователь ввел username
        telegram_id = await get_user_id_by_username(input_text, bot)
        if not telegram_id:
            await message.answer(
                "❌ Пользователь с таким username не найден. Убедитесь, что он существует и бот может его видеть (например, если пользователь ранее взаимодействовал с ботом).")
            return
    elif input_text.startswith('https://t.me/'):
        # Пользователь ввел ссылку на профиль
        username = input_text.split('/')[-1]
        telegram_id = await get_user_id_by_username(username, bot)
        if not telegram_id:
            await message.answer("❌ Пользователь по этой ссылке не найден.")
            return
    elif input_text.isdigit():
        # Пользователь ввел цифровой ID
        telegram_id = int(input_text)
    else:
        await message.answer("❌ Неверный формат. Введите ID, username (с @) или ссылку на профиль.")
        return

    async with async_session() as session:
        try:
            result = await session.execute(
                select(User).where(User.telegram_id == telegram_id)
            )
            user = result.scalar_one_or_none()
            try:
                chat_member = await bot.get_chat_member(telegram_id, telegram_id)
                full_name = getattr(chat_member.user, 'full_name', f"Admin {telegram_id}")
                username = getattr(chat_member.user, 'username', None)
            except Exception as e:
                print(f"Не удалось получить данные пользователя: {e}")
                full_name = f"Admin {telegram_id}"
                username = None

            if user:
                user.is_admin = True
                user.full_name = full_name or user.full_name
                user.username = username or user.username
            else:
                user = User(
                    telegram_id=telegram_id,
                    full_name=full_name,
                    username=username,
                    is_admin=True
                )
                session.add(user)

            await session.commit()
            await message.answer(
                f"✅ Пользователь <b>{user.full_name}</b> (ID: <code>{telegram_id}</code>) теперь администратор.",
                reply_markup=admin_menu_kb(),
                parse_mode="HTML"
            )
        except Exception as e:
            await session.rollback()
            await message.answer(
                f"❌ Ошибка при сохранении: {str(e)}",
                reply_markup=admin_menu_kb()
            )

    await state.clear()


@router.message(F.text == "Добавить тему", IsAdminFilter())
async def add_topic_start(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Введите название темы:", reply_markup=cancel_kb())
    await state.set_state(AddTopic.waiting_for_title)


@router.message(AddTopic.waiting_for_title)
async def process_title(message: Message, state: FSMContext):
    await state.update_data(title=message.text)
    await message.answer("Введите описание темы:")
    await state.set_state(AddTopic.waiting_for_description)


@router.message(AddTopic.waiting_for_description)
async def process_description(message: Message, state: FSMContext):
    await state.update_data(description=message.text)
    await message.answer(
        "Введите сложность темы (легкая/средняя/сложная):\n"
        "Доступные варианты: easy, medium, hard",
        reply_markup=cancel_kb()
    )
    await state.set_state(AddTopic.waiting_for_difficulty)


@router.message(AddTopic.waiting_for_difficulty)
async def process_difficulty(message: Message, state: FSMContext):
    difficulty = message.text.strip().lower()
    if difficulty not in ALLOWED_DIFFICULTIES:
        await message.answer(
            f"❌ Неверная сложность. Допустимые варианты: {', '.join(ALLOWED_DIFFICULTIES)}.\n"
            "Попробуйте еще раз:"
        )
        return

    await state.update_data(difficulty=difficulty)
    await message.answer(
        "Введите направление темы:\n"
        f"Доступные варианты: {', '.join(ALLOWED_DIRECTIONS)}",
        reply_markup=cancel_kb()
    )
    await state.set_state(AddTopic.waiting_for_direction)


@router.message(AddTopic.waiting_for_direction)
async def process_direction(message: Message, state: FSMContext):
    direction = message.text.strip().lower()
    if direction not in ALLOWED_DIRECTIONS:
        await message.answer(
            f"❌ Неверное направление. Допустимые варианты: {', '.join(ALLOWED_DIRECTIONS)}.\n"
            "Попробуйте еще раз:"
        )
        return

    await state.update_data(direction=direction)
    await message.answer(
        f"Введите 8 весов (числа от 0.0 до 1.0) через пробел.\n"
        f"Например: 1.0 0.0 0.2 0.5 0.8 0.0 0.3 0.4\n"
        f"Значения будут округлены до двух знаков после запятой.",
        reply_markup=cancel_kb()
    )
    await state.set_state(AddTopic.waiting_for_weights)


@router.message(AddTopic.waiting_for_weights)
async def process_weights(message: Message, state: FSMContext):
    parts = message.text.strip().split()
    if len(parts) != REQUIRED_WEIGHTS_COUNT:   # 8
        await message.answer(f"❌ Нужно ровно {REQUIRED_WEIGHTS_COUNT} чисел через пробел.")
        return
    try:
        weights = [float(x) for x in parts]
    except ValueError:
        await message.answer("❌ Все значения должны быть числами.")
        return
    if any(w < 0.0 or w > 1.0 for w in weights):
        await message.answer("❌ Числа должны быть от 0.0 до 1.0.")
        return

    data = await state.get_data()
    async with async_session() as session:
        topic = Topic(
            title=data['title'],
            description=data['description'],
            difficulty=data['difficulty'],
            direction=data['direction'],
            weights=weights
        )
        session.add(topic)
        await session.commit()

    await message.answer(
        "✅ Тема успешно добавлена!\n"
        f"Название: {data['title']}\n"
        f"Сложность: {data['difficulty'].capitalize()}\n"
        f"Направление: {data['direction']}\n"
        f"Веса: {', '.join(f'{w:.2f}' for w in weights)}",
        reply_markup=admin_menu_kb()
    )
    await state.clear()


@router.message(F.text == "Список тем", IsAdminFilter())
async def export_topics_excel(message: Message, state: FSMContext):
    await state.clear()
    async with async_session() as session:
        result = await session.execute(select(Topic))
        topics = result.scalars().all()

    if not topics:
        await message.answer("Тем пока нет.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Темы курсовых"
    ws.append(["Название", "Описание", "Сложность", "Направление", "Веса"])

    for topic in topics:
        weights_str = ", ".join(f"{float(w):.2f}" if w is not None else "None" for w in topic.weights)
        ws.append([
            topic.title,
            topic.description,
            topic.difficulty.capitalize(),
            topic.direction,
            weights_str
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await message.answer_document(
        BufferedInputFile(buffer.read(), filename="topics.xlsx"),
        caption="📋 Список всех тем"
    )


@router.message(F.text == "В главное меню", IsAdminFilter())
async def back_to_main_menu(message: Message):
    await message.answer("Возвращаемся в главное меню.", reply_markup=main_menu_kb())

@router.message(F.text == "Удалить все темы", IsAdminFilter())
async def delete_all_topics(message: Message):
    async with async_session() as session:
        await session.execute(delete(Topic))
        await session.commit()
    await message.answer("❌ Все темы были удалены.")


@router.message(F.text == "/debug_weights", IsAdminFilter())
async def debug_weights(message: Message):
    async with async_session() as session:
        try:
            topic = (await session.execute(
                select(Topic).limit(1)
            )).scalars().first()

            if not topic:
                return await message.answer("В базе нет ни одной темы")

            if not hasattr(topic, 'weights'):
                return await message.answer("У тем нет поля weights")

            debug_info = [
                "🔍 Отладочная информация:",
                f"ID темы: {topic.id}",
                f"Тип weights: {type(topic.weights)}",
            ]

            if topic.weights:
                debug_info.extend([
                    f"Тип элемента: {type(topic.weights[0])}",
                    f"Пример значения: {topic.weights[:3]}... (всего {len(topic.weights)})",
                    f"SQL для проверки: SELECT weights FROM topics WHERE id = {topic.id};"
                ])
            else:
                debug_info.append("weights: пустой массив")

            await message.answer("\n".join(debug_info))

        except Exception as e:
            await message.answer(f"Ошибка при проверке: {str(e)}")